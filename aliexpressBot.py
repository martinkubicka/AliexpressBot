from bs4 import BeautifulSoup
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains

def ctor():
    driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())
    driver.set_window_size(1200, 1100)
    driver.set_window_position(-700, 2000)
    excelfile = "output.xlsx"
    vk = openpyxl.load_workbook(excelfile)
    std = vk["program"]
    vk.remove(std)
    vk.create_sheet("program", 0)
    sh = vk["program"]
    sh.cell(row=1, column=1).value = "Aliexpress Link"
    sh.cell(row=1, column=2).value = "Justia Brand"
    sh.cell(row=1, column=3).value = "Vero Title"
    sh.cell(row=1, column=4).value = "Vero Brand"
    vk.save(excelfile)
    return driver, vk, sh

def dtor(driver):
    driver.quit()

def getJustiaTrademarks(brand, driver):
    driver.get(("https://trademarks.justia.com/search?q={}").format(brand))
    try:
        trademark = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//h4[@class='has-no-margin heading-5']"))).text
    except:
        trademark = "-"
    return trademark

def yabelleLogin(driver):
    action = ActionChains(driver)
    driver.get("https://yaballe.com/login")
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//a[@href='#login']"))).click()
    time.sleep(0.5)
    email = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='login_email']")))
    action.send_keys_to_element(email, "testamind@gmail.com").perform()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='password']"))).send_keys("lA029921026")
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//button[@id='login_site']"))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[text()='Lister']")))

def yabelleSearch(driver, link):
    driver.get("https://yaballe.com/#/autolister")
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//input[@id='asinTextBox']"))).send_keys("https://www.aliexpress.com" + link)
    time.sleep(0.2)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//button[@id='btn-amazon']"))).click()
    WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//h3[@class='title-hero']")))
    title_vero = "0"
    brand_vero = "0"
    for vero in driver.find_elements_by_xpath("//li[@ng-repeat='(key, value) in product.vero_warnings']"):
        if "brand" in vero.text:
            brand_vero = vero.text.split("brand: ")[1]
        elif "title" in vero.text:
            title_vero = vero.text.split("title: ")[1]
    return title_vero, brand_vero

def writeToExcel(link, trademark, title_vero, brand_vero, vk, sh):
    sh.cell(row=sh.max_row+1, column=1).value = link
    sh.cell(row=sh.max_row, column=2).value = trademark
    sh.cell(row=sh.max_row, column=3).value = title_vero
    sh.cell(row=sh.max_row, column=4).value = brand_vero
    vk.save("output.xlsx")

def getLinksAliexpress(keywords, driver, vk, sh):
    for i in keywords:
        print("Searching keyword " + i)
        driver.get(("https://www.aliexpress.com/wholesale?SearchText={}").format(i))
        time.sleep(0.5)
        y = 100
        for j in range(30):
            driver.execute_script("window.scrollTo(0, {})".format(y))
            y += 200
        time.sleep(1)
        soup = BeautifulSoup(driver.page_source, "html.parser")
        for item in soup.find_all("a", class_="_3t7zg _2f4Ho"):
            brand = getBrandNameAliexpress(item["href"], driver)
            if brand != "-":
                trademark = getJustiaTrademarks(brand, driver)
            else:
                trademark = "-"
            vero = yabelleSearch(driver, item["href"])
            writeToExcel("https://www.aliexpress.com" + item["href"], trademark, vero[0], vero[1], vk, sh)

def getBrandNameAliexpress(item, driver):
    driver.get("https://www.aliexpress.com" + item)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    brand = ""
    for i in range(int(soup.text.find('{"@type":"Brand","name":')) + len('{"@type":"Brand","name":"'), len(soup.text)):
        if "}" in soup.text[i]:
            break
        elif '"' in soup.text[i]:
            continue
        elif "/" in soup.text[i]:
            brand = "-"
            break
        else:
            brand = brand + soup.text[i]
    return brand

def getKeywords():
    keywords = []
    file = open("input.txt", "r")
    for i in file.readlines():
        keywords.append(i.replace("\n", ""))
    return keywords

def main():
    print("Successfully started")
    keywords = getKeywords()
    ctorvalues = ctor()
    yabelleLogin(ctorvalues[0])
    getLinksAliexpress(keywords, ctorvalues[0], ctorvalues[1], ctorvalues[2])
    dtor(ctorvalues[0])
    print("Program ended")
    input("Press enter to continue..")

if __name__ == '__main__':
    main()
    #comments